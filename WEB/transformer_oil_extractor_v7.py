#!/usr/bin/env python3
"""
Transformer Oil Test Report Extractor  —  UNIFIED
===================================================
Supports two report formats detected automatically:
  • TRU-FIL  : IS-standard references, Equipment Designation / CSS / TR# layout
  • SGS/CPRI : Unicode subscripts (H₂, CH₄ …), Owner / Installation Location layout

Reads one or more PDFs per session and appends every extracted field into a
single Excel file.  Re-run at any time — new rows are appended, old data is
preserved.

Usage:
    python3 transformer_oil_extractor_unified.py

Dependencies (auto-installed):
    pip install pdfplumber openpyxl
"""

import os, re, sys
from tkinter import Tk
from tkinter.filedialog import askopenfilenames

EXCEL_FILE = "transformer_oil_data.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Column definitions  (order here = order in Excel)
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = [
    # --- Identity / location ---
    "Format",
    "Owner",
    "CSS Name",
    "Installation Location",
    "Equipment Designation",
    "Equipment Type",
    "Transformer No. / Serial No.",
    "Manufacturer",
    "Manufacturer Sl No",
    "Rating (KVA)",
    "Voltage Class (KV)",
    "Voltage Ratio",
    "Cooling",
    "Manufacturing Year",
    "Type of Oil / Product Name",
    # --- Dates / report metadata ---
    "Report Number",
    "Sample ID",
    "Report Date",
    "Sampling Date",
    "Weather Condition",
    # --- Oil Screening Tests (OST) ---
    "BDV (KV)",
    "Water Content (ppm / mg/kg)",
    "Color (ASTM)",
    "Density g/cm³",
    "Sp.Res 27°C (×10¹² Ω·cm)",
    "Sp.Res 90°C (×10¹² Ω·cm)",
    "DDF 27°C (tan δ)",
    "DDF 90°C (tan δ)",
    "IFT (N/m)",
    "Neutralization (mgKOH/g)",
    "Sediment / Sludge (%)",
    "Flash Point (°C)",
    "OQI",
    # --- Dissolved Gas Analysis (DGA) ---
    "H₂  (ppm / µl·l⁻¹)",
    "O₂  (ppm / µl·l⁻¹)",
    "N₂  (ppm / µl·l⁻¹)",
    "CO  (ppm / µl·l⁻¹)",
    "CH₄ (ppm / µl·l⁻¹)",
    "CO₂ (ppm / µl·l⁻¹)",
    "C₂H₄ (ppm / µl·l⁻¹)",
    "C₂H₆ (ppm / µl·l⁻¹)",
    "C₂H₂ (ppm / µl·l⁻¹)",
    "C₃H₆ (ppm / µl·l⁻¹)",
    "C₃H₈ (ppm / µl·l⁻¹)",
    "TDCG (mg/kg or ppm)",
    "TGC (v/v %)",
    "TDCG/TGC (%)",
    # --- Recommendations ---
    "OST Recommendation",
    "DGA Recommendation",
    "Overall Recommendation",
    # --- Source ---
    "Source PDF",
]

COL_WIDTHS = {
    "Format": 10,
    "Owner": 28,
    "CSS Name": 24,
    "Installation Location": 28,
    "Equipment Designation": 28,
    "Equipment Type": 18,
    "Transformer No. / Serial No.": 20,
    "Manufacturer": 24,
    "Manufacturer Sl No": 20,
    "Rating (KVA)": 12,
    "Voltage Class (KV)": 14,
    "Voltage Ratio": 18,
    "Cooling": 12,
    "Manufacturing Year": 14,
    "Type of Oil / Product Name": 24,
    "Report Number": 20,
    "Sample ID": 16,
    "Report Date": 14,
    "Sampling Date": 14,
    "Weather Condition": 16,
    "BDV (KV)": 10,
    "Water Content (ppm / mg/kg)": 16,
    "Color (ASTM)": 10,
    "Density g/cm³": 13,
    "Sp.Res 27°C (×10¹² Ω·cm)": 18,
    "Sp.Res 90°C (×10¹² Ω·cm)": 18,
    "DDF 27°C (tan δ)": 14,
    "DDF 90°C (tan δ)": 14,
    "IFT (N/m)": 10,
    "Neutralization (mgKOH/g)": 18,
    "Sediment / Sludge (%)": 16,
    "Flash Point (°C)": 13,
    "OQI": 8,
    "H₂  (ppm / µl·l⁻¹)": 14,
    "O₂  (ppm / µl·l⁻¹)": 14,
    "N₂  (ppm / µl·l⁻¹)": 14,
    "CO  (ppm / µl·l⁻¹)": 14,
    "CH₄ (ppm / µl·l⁻¹)": 14,
    "CO₂ (ppm / µl·l⁻¹)": 14,
    "C₂H₄ (ppm / µl·l⁻¹)": 14,
    "C₂H₆ (ppm / µl·l⁻¹)": 14,
    "C₂H₂ (ppm / µl·l⁻¹)": 14,
    "C₃H₆ (ppm / µl·l⁻¹)": 14,
    "C₃H₈ (ppm / µl·l⁻¹)": 14,
    "TDCG (mg/kg or ppm)": 16,
    "TGC (v/v %)": 12,
    "TDCG/TGC (%)": 12,
    "OST Recommendation": 40,
    "DGA Recommendation": 40,
    "Overall Recommendation": 50,
    "Source PDF": 32,
}


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def create_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transformer Oil Data"

    hfill   = PatternFill("solid", start_color="1F4E79")
    hfont   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin    = Side(style="thin", color="000000")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr

    ws.row_dimensions[1].height = 46
    ws.freeze_panes = "A2"

    for col, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(h, 14)

    wb.save(EXCEL_FILE)
    print(f"  Created: {EXCEL_FILE}")


def append_row(data: dict, pdf_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not os.path.exists(EXCEL_FILE):
        create_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    next_row = ws.max_row + 1

    # Map header names → dict keys
    KEY_MAP = {
        "Format":                               "fmt",
        "Owner":                                "owner",
        "CSS Name":                             "css_name",
        "Installation Location":                "installation_location",
        "Equipment Designation":                "equipment_designation",
        "Equipment Type":                       "equipment_type",
        "Transformer No. / Serial No.":         "transformer_no",
        "Manufacturer":                         "manufacturer",
        "Manufacturer Sl No":                   "manufacturer_slno",
        "Rating (KVA)":                         "rating",
        "Voltage Class (KV)":                   "voltage_class",
        "Voltage Ratio":                        "voltage_ratio",
        "Cooling":                              "cooling",
        "Manufacturing Year":                   "manufacturing_year",
        "Type of Oil / Product Name":           "oil_type",
        "Report Number":                        "report_no",
        "Sample ID":                            "sample_id",
        "Report Date":                          "report_date",
        "Sampling Date":                        "sampling_date",
        "Weather Condition":                    "weather_condition",
        "BDV (KV)":                             "bdv",
        "Water Content (ppm / mg/kg)":          "water",
        "Color (ASTM)":                         "color",
        "Density g/cm³":                        "density",
        "Sp.Res 27°C (×10¹² Ω·cm)":            "sp_res_27",
        "Sp.Res 90°C (×10¹² Ω·cm)":            "sp_res_90",
        "DDF 27°C (tan δ)":                     "ddf_27",
        "DDF 90°C (tan δ)":                     "ddf_90",
        "IFT (N/m)":                            "ift",
        "Neutralization (mgKOH/g)":             "neutralization",
        "Sediment / Sludge (%)":                "sediment",
        "Flash Point (°C)":                     "flash",
        "OQI":                                  "oqi",
        "H₂  (ppm / µl·l⁻¹)":                 "h2",
        "O₂  (ppm / µl·l⁻¹)":                 "o2",
        "N₂  (ppm / µl·l⁻¹)":                 "n2",
        "CO  (ppm / µl·l⁻¹)":                  "co",
        "CH₄ (ppm / µl·l⁻¹)":                  "ch4",
        "CO₂ (ppm / µl·l⁻¹)":                  "co2",
        "C₂H₄ (ppm / µl·l⁻¹)":                "c2h4",
        "C₂H₆ (ppm / µl·l⁻¹)":                "c2h6",
        "C₂H₂ (ppm / µl·l⁻¹)":                "c2h2",
        "C₃H₆ (ppm / µl·l⁻¹)":                "c3h6",
        "C₃H₈ (ppm / µl·l⁻¹)":                "c3h8",
        "TDCG (mg/kg or ppm)":                  "tdcg",
        "TGC (v/v %)":                          "tgc",
        "TDCG/TGC (%)":                         "tdcg_ratio",
        "OST Recommendation":                   "ost_recommendation",
        "DGA Recommendation":                   "dga_recommendation",
        "Overall Recommendation":               "recommendation",
        "Source PDF":                           "__pdf__",
    }

    thin   = Side(style="thin",  color="CCCCCC")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill   = PatternFill("solid", start_color="EBF3FB") if next_row % 2 == 0 else None
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(HEADERS, 1):
        key = KEY_MAP.get(header, "")
        val = os.path.basename(pdf_path) if key == "__pdf__" else data.get(key, "")
        c = ws.cell(row=next_row, column=col, value=val)
        c.border = bdr
        c.alignment = center
        c.font = Font(name="Calibri", size=10)
        if fill:
            c.fill = fill

    ws.row_dimensions[next_row].height = 20
    wb.save(EXCEL_FILE)


# ─────────────────────────────────────────────────────────────────────────────
# Regex helper
# ─────────────────────────────────────────────────────────────────────────────

def ev(text, pattern, default="ND", flags=re.IGNORECASE | re.DOTALL):
    """Extract first captured group; return default if no match."""
    try:
        m = re.search(pattern, text, flags)
        return m.group(1).strip() if m else default
    except Exception:
        return default


# ─────────────────────────────────────────────────────────────────────────────
# Format detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(text: str) -> str:
    """
    Returns 'TRUFIL' or 'SGS'.
    TRU-FIL reports reference IS standards like "IS 6792", "IS 9434".
    SGS/CPRI reports have 'Owner' and Unicode subscript chemicals.
    """
    if re.search(r"IS\s+6792|IS\s+9434|IS\s+6103", text):
        return "TRUFIL"
    if re.search(r"Owner\s+\S|H₂|µl/l|Installation Location", text):
        return "SGS"
    # Fallback: try both and pick whichever hits more fields
    return "TRUFIL"


# ─────────────────────────────────────────────────────────────────────────────
# TRU-FIL parser  (v1 – v3 logic, unified)
# ─────────────────────────────────────────────────────────────────────────────

def parse_trufil(full: str) -> dict:
    d = {"fmt": "TRU-FIL"}

    # Equipment Designation → CSS Name + Transformer No.
    m = re.search(r"Equipment Designation\s+([^\n]+)", full, re.IGNORECASE)
    eq = m.group(1).strip() if m else ""

    # Try "CSS NAME / TR#123" pattern first, then generic two-part split
    m2 = re.search(r"(.*?)\s*/\s*(TR[#\-\d\w]+)", eq, re.IGNORECASE)
    if not m2:
        m2 = re.search(r"(.*?)\s*/\s*([A-Z]{2}[#\-]?\d+)", eq)

    if m2:
        d["css_name"]       = re.sub(r"S/[Ss]", "", m2.group(1)).strip()
        d["transformer_no"] = m2.group(2).strip()
    else:
        d["css_name"]       = eq
        d["transformer_no"] = ""

    d["equipment_designation"] = eq

    d["report_date"]   = ev(full, r"Report Date\s+([\d][\d\-/]+\d)")
    d["sampling_date"] = ev(full, r"Sampling Date\s+([\d][\d\-/]+\d)")
    d["manufacturer"]  = ev(full, r"Manufacturer\s+([^\n]+)", default="")
    d["rating"]        = ev(full, r"Rating\s+([\d,]+\s*KVA)", default="")
    d["voltage_class"] = ev(full, r"Voltage Class\s+([\d]+\s*KV)", default="")
    d["oil_type"]      = ev(full, r"Insulating Fluid\s+([^\n]+)", default="")

    # ---- OST ----
    d["bdv"]           = ev(full, r"Electric Strength[^K]+KV[^I]+IS\s*6792\s+([\d.]+)")
    d["water"]         = ev(full, r"Water Content[^m]+mg.KG[^\n]+IS\s*13567\s+([\d.]+)")
    d["color"]         = ev(full, r"Visual Appearance\s*-\s*Color[^\n]+(L\s*[\d.]+)")
    d["density"]       = ev(full, r"Density[^g]+g.cm[^\n]+IS\s*1448[^\n]+([\d.]+)\s+[\d.]")
    d["sp_res_27"]     = ev(full, r"Sp\.\s*Resistance at 27[^I]+IS\s*6103\s+([\d.]+)")
    d["sp_res_90"]     = ev(full, r"Specific Resistance\s*.90[^I]+IS\s*6103\s+([\d.]+)")
    d["ddf_27"]        = ev(full, r"Dissipation Factor[^\n]+27\s*C[^I]+IS\s*6262\s+([\d.]+)")
    d["ddf_90"]        = ev(full, r"Dissipation Factor[^\n]+90\s*C[^I]+IS\s*6262\s+([\d.]+)")
    d["ift"]           = ev(full, r"Interfacial Tension[^N]+N.m[^I]+IS\s*6104\s+([\d.]+)")
    d["neutralization"]= ev(full, r"Neutralization Value[^I]+IEC\s*62021[^\n]+([\d.]+)")
    d["sediment"]      = ev(full, r"Sediment and Sludge[^A]+Annex[^\n]+([\d.]+)")
    d["flash"]         = ev(full, r"Flash Point[^I]+IS\s*1448[^\n]+([\d]+)")
    d["oqi"]           = ev(full, r"Oil Quality Index[^\-]+\-[^\-]+\-\s+([\d]+)")

    # ---- DGA (IS 9434 style) ----
    d["h2"]   = ev(full, r"Hydrogen[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["o2"]   = ev(full, r"Oxygen[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["n2"]   = ev(full, r"Nitrogen[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["co"]   = ev(full, r"Carbon Monoxide[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["ch4"]  = ev(full, r"Methane[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["co2"]  = ev(full, r"Carbon Dioxide[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["c2h4"] = ev(full, r"Ethylene[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["c2h6"] = ev(full, r"Ethane[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["c2h2"] = ev(full, r"Acetylene[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+|ND)")
    d["c3h6"] = ev(full, r"Propylene[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+|ND)")
    d["c3h8"] = ev(full, r"Propane[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+|ND)")
    d["tdcg"] = ev(full, r"Total Dissolved Combustible[^p]+ppm[^\n]+IS\s*9434\s+([\d.]+)")
    d["tgc"]  = ev(full, r"Total Gas Content[^v]+v.v[^\n]+IS\s*9434\s+([\d.]+)")

    # ---- Recommendations ----
    rec = re.search(r"Overall Recommendations?\s*:\s*([^\n]+)", full, re.IGNORECASE)
    d["recommendation"] = rec.group(1).strip() if rec else ""
    d["ost_recommendation"] = ""
    d["dga_recommendation"] = ""

    # Fields not present in TRU-FIL
    for k in ("owner", "installation_location", "equipment_type",
              "manufacturer_slno", "voltage_ratio", "cooling",
              "manufacturing_year", "weather_condition",
              "sample_id", "report_no", "tdcg_ratio"):
        d.setdefault(k, "")

    return d


# ─────────────────────────────────────────────────────────────────────────────
# SGS / CPRI parser  (v4 – v6 logic, unified)
# ─────────────────────────────────────────────────────────────────────────────

def parse_sgs(full: str) -> dict:
    d = {"fmt": "SGS/CPRI"}

    # ---- Identity ----
    d["owner"]                  = ev(full, r"Owner\s+([^\n]+)")
    d["installation_location"]  = ev(full, r"Installation Location\s+([^\n]+)")
    d["equipment_designation"]  = ev(full, r"Equipment Designation\s+([^\n]+)")
    d["equipment_type"]         = ev(full, r"Equipment Type\s+([^\n]+)")
    d["transformer_no"]         = ev(full, r"(?:Equipment Serial No|Manufacturer'?s?\s+Sl\.?\s*No\.?)\s*[:\s]+([^\n]+)", default="")
    d["manufacturer_slno"]      = ev(full, r"Manufacturer'?s?\s+Sl\.?\s*No\.\s+([^\n]+)")
    d["manufacturer"]           = ev(full, r"Equipment Make\s*:\s*([^\n]+)", default="")
    d["rating"]                 = ev(full, r"Power Rating\s*:\s*([^\n]+)", default="")
    d["voltage_class"]          = ev(full, r"Voltage Rating\s*:\s*([^\n]+)", default="")
    d["voltage_ratio"]          = ev(full, r"Voltage Ratio[^\n]*?\s+([0-9 /]+VOLT)", default="")
    d["cooling"]                = ev(full, r"Cooling\s+([^\n]+)")
    d["manufacturing_year"]     = ev(full, r"Manufacturing Year\s+([^\n]+)")
    d["oil_type"]               = ev(full, r"Product Name\s*:\s*([^\n]+)", default="")
    d["weather_condition"]      = ev(full, r"Weather [Cc]ondition\s+([^\n]+)")
    d["sample_id"]              = ev(full, r"Our Sample ID\s+([A-Z0-9\-]+)")
    d["report_no"]              = ev(full, r"Oil Test Report\s*[-–]\s*([A-Z0-9/\-]+)")
    d["report_date"]            = ev(full, r"Analysis Date\s*[:\s]+([0-9/\-]+)")
    d["sampling_date"]          = ev(full, r"Sampling Date\s*[:\s]+([0-9/\-]+)")

    # css_name not present; fill from equipment_designation
    d["css_name"] = d["equipment_designation"]

    # ---- OST — two sub-formats seen across v4/v5/v6 ----
    # v6 uses "BDV" label + limit "30"; v4/v5 use "Breakdown Voltage" + "Good"
    d["bdv"] = (
        ev(full, r"Electric Strength \(BDV\)[^\n]*?([0-9.]+)\s+30") or
        ev(full, r"Electric Strength \(Breakdown Voltage\)[^\n]*?([0-9.]+)\s+(?:Good|30)", default="ND")
    )
    d["water"] = (
        ev(full, r"Water Content By Karl Fischer Method[^\n]*?([0-9.]+)\s+40") or
        ev(full, r"Water Content[^\n]*?mg/kg[^\n]*?([0-9.]+)\s+(?:Good|40)", default="ND")
    )
    d["neutralization"] = ev(full, r"Neutralization Value[^\n]*?([0-9.]+)\s+0\.3")
    if d["neutralization"] == "ND":
        d["neutralization"] = ev(full, r"Neutralization Value[^\n]*?([0-9.]+)\s+Good")

    d["ift"] = ev(full, r"Interfacial Tension[^\n]*?([0-9.]+)\s+0\.020")
    if d["ift"] == "ND":
        d["ift"] = ev(full, r"Interfacial Tension[^\n]*?([0-9.]+)\s+Good")

    d["ddf_90"] = ev(full, r"Dielectric Dissipation Factor[^\n]*?90[°o]?C[^\n]*?([0-9.]+)\s+(?:Good|0\.[0-9]+)")
    d["ddf_27"] = ev(full, r"Dielectric Dissipation Factor[^\n]*?(?:27[°o]?C|RT)[^\n]*?([0-9.]+)")
    d["sp_res_90"] = ev(full, r"Specific Resistance[^\n]*?90[°o]?C[^\n]*?([0-9.]+)\s+(?:Good|[0-9.]+)")
    d["sp_res_27"] = ev(full, r"Specific Resistance[^\n]*?(?:27[°o]?C|RT)[^\n]*?([0-9.]+)")
    d["sediment"] = ev(full, r"Sediment[^\n]*?(Not Detected|[0-9.]+)")
    d["flash"]    = ev(full, r"Flash Point[^\n]*?([0-9]+)")
    d["density"]  = (
        ev(full, r"Density\s*@[^\n]*?([0-9.]+)\s+0\.890") or
        ev(full, r"Density[^\n]*?g/ml\s*([0-9.]+)", default="ND")
    )

    # ---- DGA (Unicode subscript style) ----
    #   Each row:  Gas name (symbol)  value   limit
    #   We match the unicode symbol variants flexibly.

    def dga(pattern):
        return ev(full, pattern, default="ND")

    d["h2"]   = dga(r"Hydrogen\s*\(H[₂2]\)[^\n]*?([0-9.]+)\s+50")
    d["o2"]   = dga(r"Oxygen\s*\(O[₂2]\)[^\n]*?([0-9.]+)\s+NS")
    d["n2"]   = dga(r"Nitrogen\s*\(N[₂2]\)[^\n]*?([0-9.]+)\s+NS")
    d["co"]   = dga(r"Carbon Monoxide\s*\(CO\)[^\n]*?([0-9.]+)\s+400")
    d["ch4"]  = dga(r"Methane\s*\(CH[₄4]\)[^\n]*?([0-9.]+)\s+30")
    d["co2"]  = dga(r"Carbon Dioxide\s*\(CO[₂2]\)[^\n]*?([0-9.]+)\s+3800")
    d["c2h4"] = dga(r"Ethylene\s*\(C[₂2]H[₄4]\)[^\n]*?([0-9.]+)\s+60")
    d["c2h6"] = dga(r"Ethane\s*\(C[₂2]H[₆6]\)[^\n]*?([0-9.]+)\s+20")
    d["c2h2"] = dga(r"Acetylene\s*\(C[₂2]H[₂2]\)[^\n]*?(ND|[0-9.]+)\s+2")
    d["c3h6"] = dga(r"Propylene\s*\(C[₃3]H[₆6]\)[^\n]*?(ND|[0-9.]+)\s+NS")
    d["c3h8"] = dga(r"Propane\s*\(C[₃3]H[₈8]\)[^\n]*?(ND|[0-9.]+)\s+NS")
    d["tdcg"] = dga(r"Total Dissolved Combustible[^\n]*?([0-9.]+)\s+NS")
    d["tgc"]  = dga(r"Total Gas Content\s*\(TGC\)[^\n]*?([0-9.]+)\s+NS")
    d["tdcg_ratio"] = dga(r"TDCG/TGC[^\n]*?([0-9.]+)\s+8%")

    # ---- Recommendations ----
    d["ost_recommendation"] = ev(full, r"OST Test Recommendation\s+(.*?)\n#")
    d["dga_recommendation"] = ev(full, r"DGA Test Recommendation\s+(.*?)\n#")
    overall = ev(full, r"Overall Recommendations?\s*:\s*(.*?)\nOil Test Report")
    if overall == "ND":
        overall = ev(full, r"Overall Recommendations?\s*:\s*([^\n]+)", default="")
    d["recommendation"] = overall

    # Not present in SGS
    for k in ("color", "oqi"):
        d.setdefault(k, "")

    return d


# ─────────────────────────────────────────────────────────────────────────────
# Top-level PDF parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_pdf(path: str) -> dict:
    import pdfplumber
    full = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full += t + "\n"

    fmt = detect_format(full)
    if fmt == "TRUFIL":
        d = parse_trufil(full)
    else:
        d = parse_sgs(full)

    # Guarantee every key exists
    all_keys = [
        "fmt", "owner", "css_name", "installation_location",
        "equipment_designation", "equipment_type", "transformer_no",
        "manufacturer", "manufacturer_slno", "rating", "voltage_class",
        "voltage_ratio", "cooling", "manufacturing_year", "oil_type",
        "report_no", "sample_id", "report_date", "sampling_date",
        "weather_condition", "bdv", "water", "color", "density",
        "sp_res_27", "sp_res_90", "ddf_27", "ddf_90", "ift",
        "neutralization", "sediment", "flash", "oqi",
        "h2", "o2", "n2", "co", "ch4", "co2",
        "c2h4", "c2h6", "c2h2", "c3h6", "c3h8",
        "tdcg", "tgc", "tdcg_ratio",
        "ost_recommendation", "dga_recommendation", "recommendation",
    ]
    for k in all_keys:
        d.setdefault(k, "")

    return d


# ─────────────────────────────────────────────────────────────────────────────
# Main loop
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 68)
    print("  Transformer Oil Report Extractor  |  UNIFIED (TRU-FIL + SGS/CPRI)")
    print(f"  Output : {EXCEL_FILE}")
    print("  Select one or more PDFs at once; type 'exit' to quit.")
    print("=" * 68)

    if not os.path.exists(EXCEL_FILE):
        create_excel()
    else:
        import openpyxl
        rows = openpyxl.load_workbook(EXCEL_FILE).active.max_row - 1
        print(f"  Existing file found  —  {rows} data row(s) already present.")

    while True:
        print()
        choice = input("  Press ENTER to select PDF(s)  or  type 'exit': ").strip().lower()

        if choice == "exit":
            print(f"\n  Done. Data saved to: {EXCEL_FILE}\n")
            break

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        pdf_paths = askopenfilenames(
            title="Select Transformer Oil PDF report(s)",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
            parent=root,
        )
        root.destroy()

        if not pdf_paths:
            print("  No files selected.")
            continue

        ok = err = 0
        for pdf_path in pdf_paths:
            if not pdf_path.lower().endswith(".pdf"):
                print(f"  Skipping non-PDF: {os.path.basename(pdf_path)}")
                continue

            try:
                fname = os.path.basename(pdf_path)
                print(f"\n  [{fname}]  Parsing ...", end=" ", flush=True)
                d = parse_pdf(pdf_path)
                append_row(d, pdf_path)
                print("OK")
                print(f"    Format     : {d['fmt']}")
                print(f"    CSS/Owner  : {d.get('css_name') or d.get('owner') or '--'}")
                print(f"    Transf. No : {d.get('transformer_no') or '--'}")
                print(f"    Oil Type   : {d.get('oil_type') or '--'}")
                print(f"    BDV (KV)   : {d.get('bdv') or '--'}")
                print(f"    Water      : {d.get('water') or '--'}")
                print(f"    OQI        : {d.get('oqi') or '--'}")
                print(f"    H₂/CO/C₂H₂: {d.get('h2','--')} / {d.get('co','--')} / {d.get('c2h2','--')}")
                ok += 1
            except Exception as exc:
                print(f"  ERROR — {os.path.basename(pdf_path)}: {exc}")
                err += 1

        print(f"\n  Session complete: {ok} row(s) appended, {err} error(s).")
        print(f"  File: {os.path.abspath(EXCEL_FILE)}")


if __name__ == "__main__":
    try:
        import pdfplumber, openpyxl
    except ImportError:
        print("  Installing required packages (pdfplumber, openpyxl) ...")
        os.system(f'"{sys.executable}" -m pip install pdfplumber openpyxl --quiet')
        import pdfplumber, openpyxl
    main()