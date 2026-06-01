#!/usr/bin/env python3
"""
Transformer Oil Test Report Extractor
======================================
Reads TRU-FIL PDF oil test reports and appends extracted data into an Excel file.
Keeps looping until you type 'exit'.

Usage:
    python3 transformer_oil_extractor.py

Dependencies (auto-installed if missing):
    pip install pdfplumber openpyxl
"""

import os, re, sys
from tkinter import Tk
from tkinter.filedialog import askopenfilename

EXCEL_FILE = "transformer_oil_data.xlsx"

HEADERS = [
    "CSS Name", "Transformer No.", "Report Date", "Sampling Date",
    "Manufacturer", "Rating (KVA)", "Voltage Class (KV)", "Type of Oil",
    "BDV (KV)", "Water Content (ppm)", "Color (ASTM)", "Density g/cm3",
    "Sp.Res 27C (x10^12 Ohm-cm)", "Sp.Res 90C (x10^12 Ohm-cm)",
    "DDF 27C (tand)", "DDF 90C (tand)", "IFT N/m",
    "Neutralization mgKOH/g", "Sediment %", "Flash Point C", "OQI",
    "H2 ppm", "O2 ppm", "N2 ppm", "CO ppm", "CH4 ppm", "CO2 ppm",
    "C2H4 ppm", "C2H6 ppm", "C2H2 ppm", "C3H6 ppm", "C3H8 ppm",
    "TDCG mg/kg", "TGC v/v%", "Overall Recommendation", "Source PDF",
]

def create_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transformer Oil Data"
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr
    ws.row_dimensions[1].height = 42
    widths = [22,14,14,14,24,12,14,22,10,14,10,14,20,20,10,10,12,18,12,12,10,
              10,12,14,10,10,10,10,10,10,10,10,14,12,45,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(EXCEL_FILE)

def ev(text, pattern, default="ND"):
    try:
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else default
    except Exception:
        return default

def parse_pdf(path):
    import pdfplumber
    full = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: full += t + "\n"

    d = {}

    # CSS Name + Transformer number from Equipment Designation line
    m = re.search(r"Equipment Designation\s+([^\n]+)", full, re.IGNORECASE)
    eq = m.group(1).strip() if m else ""
    m2 = re.search(r"(.*?)\s*/\s*([A-Z]{2}[#\-]?\d+)", eq)

    if m2:
        d["css_name"] = (
            m2.group(1)
            .replace("S/s", "")
            .replace("S/S", "")
            .strip()
        )

        d["transformer_no"] = m2.group(2).strip()

    else:
        d["css_name"] = eq
        d["transformer_no"] = ""

    d["report_date"]    = ev(full, r"Report Date\s+([\d][\d\-/]+\d)")
    d["sampling_date"]  = ev(full, r"Sampling Date\s+([\d][\d\-/]+\d)")
    d["manufacturer"]   = ev(full, r"Manufacturer\s+([^\n]+)", default="")
    d["rating"]         = ev(full, r"Rating\s+([\d]+\s*KVA)", default="")
    d["voltage_class"]  = ev(full, r"Voltage Class\s+([\d]+\s*KV)", default="")
    d["oil_type"]       = ev(full, r"Insulating Fluid\s+([^\n]+)", default="")

    # Oil screening tests
    d["bdv"]            = ev(full, r"Electric Strength[^K]+KV[^I]+IS 6792\s+([\d.]+)")
    d["water"]          = ev(full, r"Water Content[^m]+mg.KG[^\n]+IS 13567\s+([\d.]+)")
    d["color"]          = ev(full, r"Visual Appearance - Color[^\n]+(L\s*[\d.]+)")
    d["density"]        = ev(full, r"Density[^g]+g.cm[^\n]+IS 1448[^\n]+([\d.]+)\s+[\d.]")
    d["sp_res_27"]      = ev(full, r"Sp\. Resistance at 27[^I]+IS 6103\s+([\d.]+)")
    d["sp_res_90"]      = ev(full, r"Specific Resistance .90[^I]+IS 6103\s+([\d.]+)")
    d["ddf_27"]         = ev(full, r"Dissipation Factor[^\n]+27 C[^I]+IS 6262\s+([\d.]+)")
    d["ddf_90"]         = ev(full, r"Dissipation Factor[^\n]+90 C[^I]+IS 6262\s+([\d.]+)")
    d["ift"]            = ev(full, r"Interfacial Tension[^N]+N.m[^I]+IS 6104\s+([\d.]+)")
    d["neutralization"] = ev(full, r"Neutralization Value[^I]+IEC 62021[^\n]+([\d.]+)")
    d["sediment"]       = ev(full, r"Sediment and Sludge[^A]+Annex[^\n]+([\d.]+)")
    d["flash"]          = ev(full, r"Flash Point[^I]+IS 1448[^\n]+([\d]+)")
    d["oqi"]            = ev(full, r"Oil Quality Index[^\-]+\-[^\-]+\-\s+([\d]+)")

    # DGA gases
    d["h2"]             = ev(full, r"Hydrogen[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["o2"]             = ev(full, r"Oxygen[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["n2"]             = ev(full, r"Nitrogen[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["co"]             = ev(full, r"Carbon Monoxide[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["ch4"]            = ev(full, r"Methane[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["co2"]            = ev(full, r"Carbon Dioxide[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["c2h4"]           = ev(full, r"Ethylene[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["c2h6"]           = ev(full, r"Ethane[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["c2h2"]           = ev(full, r"Acetylene[^p]+ppm[^\n]+IS 9434\s+([\d.]+|ND)")
    d["c3h6"]           = ev(full, r"Propylene[^p]+ppm[^\n]+IS 9434\s+([\d.]+|ND)")
    d["c3h8"]           = ev(full, r"Propane[^p]+ppm[^\n]+IS 9434\s+([\d.]+|ND)")
    d["tdcg"]           = ev(full, r"Total Dissolved Combustible[^p]+ppm[^\n]+IS 9434\s+([\d.]+)")
    d["tgc"]            = ev(full, r"Total Gas Content[^v]+v.v[^\n]+IS 9434\s+([\d.]+)")

    rec = re.search(r"Overall Recommendations?:\s*([^\n]+)", full, re.IGNORECASE)
    d["recommendation"] = rec.group(1).strip() if rec else ""
    return d

def append_row(data, pdf_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    if not os.path.exists(EXCEL_FILE):
        create_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    next_row = ws.max_row + 1
    row_data = [
        data.get("css_name",""),     data.get("transformer_no",""),
        data.get("report_date",""),  data.get("sampling_date",""),
        data.get("manufacturer",""), data.get("rating",""),
        data.get("voltage_class",""),data.get("oil_type",""),
        data.get("bdv",""),          data.get("water",""),
        data.get("color",""),        data.get("density",""),
        data.get("sp_res_27",""),    data.get("sp_res_90",""),
        data.get("ddf_27",""),       data.get("ddf_90",""),
        data.get("ift",""),          data.get("neutralization",""),
        data.get("sediment",""),     data.get("flash",""),
        data.get("oqi",""),
        data.get("h2",""),   data.get("o2",""),   data.get("n2",""),
        data.get("co",""),   data.get("ch4",""),  data.get("co2",""),
        data.get("c2h4",""), data.get("c2h6",""), data.get("c2h2",""),
        data.get("c3h6",""), data.get("c3h8",""), data.get("tdcg",""),
        data.get("tgc",""),
        data.get("recommendation",""),
        os.path.basename(pdf_path),
    ]
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", start_color="EBF3FB") if next_row % 2 == 0 else None
    center = Alignment(horizontal="center", vertical="center")
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.border = bdr
        c.alignment = center
        c.font = Font(name="Arial", size=10)
        if fill: c.fill = fill
    ws.row_dimensions[next_row].height = 18
    wb.save(EXCEL_FILE)

def main():
    print("=" * 64)
    print("  Transformer Oil Test Report Extractor  |  TRU-FIL format")
    print(f"  Output: {EXCEL_FILE}")
    print("  Type  : 'exit' to quit")
    print("=" * 64)

    if not os.path.exists(EXCEL_FILE):
        create_excel()
        print("  Excel file created.")
    else:
        import openpyxl
        rows = openpyxl.load_workbook(EXCEL_FILE).active.max_row - 1
        print(f"  Existing file found with {rows} data row(s).")

    while True:
        print()
        choice = input("Press ENTER to select PDF or type 'exit': ").strip().lower()

        if choice == "exit":
            print(f"\n  Done. All data saved to: {EXCEL_FILE}")
            break

        Tk().withdraw()

        pdf_path = askopenfilename(
            title="Select Transformer Oil PDF",
            filetypes=[("PDF Files", "*.pdf")]
        )

        if not pdf_path:
            print("  No file selected.")
            continue

        if not os.path.exists(pdf_path):
            print(f"  ERROR: File not found -- {pdf_path}")
            continue

        if not pdf_path.lower().endswith(".pdf"):
            print("  ERROR: Please provide a .pdf file.")
            continue

        try:
            print(f"  Parsing {os.path.basename(pdf_path)} ...")
            d = parse_pdf(pdf_path)
            append_row(d, pdf_path)
            print(f"  OK  Row appended to {EXCEL_FILE}")
            print(f"     CSS Name       : {d.get('css_name','--')}")
            print(f"     Transformer    : {d.get('transformer_no','--')}")
            print(f"     Oil Type       : {d.get('oil_type','--')}")
            print(f"     BDV (KV)       : {d.get('bdv','--')}")
            print(f"     Water (ppm)    : {d.get('water','--')}")
            print(f"     OQI            : {d.get('oqi','--')}")
            print(f"     H2/CO/C2H2     : {d.get('h2','--')} / {d.get('co','--')} / {d.get('c2h2','--')}")
        except Exception as e:
            print(f"  ERROR: {e}")

if __name__ == "__main__":
    try:
        import pdfplumber, openpyxl
    except ImportError:
        print("  Installing required packages...")
        os.system(f'"{sys.executable}" -m pip install pdfplumber openpyxl --quiet')
        import pdfplumber, openpyxl
    main()
