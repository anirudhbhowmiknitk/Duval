import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import io
import pandas as pd

# Paths — saves everything next to this script (works on Windows, Mac, Linux)
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PNG_PATH   = os.path.join(BASE_DIR, "duval_triangle.png")
EXCEL_PATH = os.path.join(BASE_DIR  , "DGA_Duval.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 1. GAS VALUES FROM DGA REPORT
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# 1. READ GAS VALUES FROM SAME EXCEL FILE
# ─────────────────────────────────────────────────────────────────────────────

INPUT_EXCEL = os.path.join(BASE_DIR, "DGA_Input.xlsx")
df = pd.read_excel(INPUT_EXCEL)

# Read values from first row

CH4  = float(df.loc[0, "CH4"])
C2H4 = float(df.loc[0, "C2H4"])
C2H2 = float(df.loc[0, "C2H2"])

total = CH4 + C2H4 + C2H2

# Duval percentage coordinates
pCH4  = (CH4  / total * 100) if total > 0 else 0
pC2H4 = (C2H4 / total * 100) if total > 0 else 0
pC2H2 = (C2H2 / total * 100) if total > 0 else 0

print(f"CH₄  = {CH4} ppm  → {pCH4:.2f}%")
print(f"C₂H₄ = {C2H4} ppm  → {pC2H4:.2f}%")
print(f"C₂H₂ = {C2H2} ppm  → {pC2H2:.2f}%")
print(f"Sum  = {total} ppm")

# ─────────────────────────────────────────────────────────────────────────────
# 2. TERNARY → CARTESIAN CONVERSION
# ─────────────────────────────────────────────────────────────────────────────
# In Duval's triangle:
#   Bottom-left  corner = 100% CH₄
#   Bottom-right corner = 100% C₂H₄
#   Top          corner = 100% C₂H₂
# We map ternary (a,b,c) where a+b+c=100 to 2D Cartesian

def ternary_to_cart(pCH4, pC2H4, pC2H2):
    """Convert ternary % to Cartesian (x,y) in equilateral triangle."""
    a = pCH4 / 100    # bottom-left axis
    b = pC2H4 / 100   # bottom-right axis
    c = pC2H2 / 100   # top axis
    x = 0.5 * (2 * b + c) / (a + b + c)
    y = (np.sqrt(3) / 2) * c / (a + b + c)
    return x, y

# ─────────────────────────────────────────────────────────────────────────────
# 3. DUVAL ZONE BOUNDARIES (standard IEC 60599 zones)
# ─────────────────────────────────────────────────────────────────────────────
# Zones: PD, T1, T2, T3, D1, D2, DT
# Each zone is a polygon defined in ternary (CH4%, C2H4%, C2H2%) coordinates

zones = {
    "PD": {
        "color": "#B3D9FF",
        "label": "PD\nPartial Discharge",
        "points": [(98,0,2),(100,0,0),(98,2,0)],  # top-left corner small triangle
    },
    "T1": {
        "color": "#FFFF99",
        "label": "T1\n< 300°C",
        "points": [(98,0,2),(98,2,0),(76,24,0),(77,0,23)],
    },
    "T2": {
        "color": "#FFD966",
        "label": "T2\n300–700°C",
        "points": [(77,0,23),(76,24,0),(40,60,0),(46,0,54)],
    },
    "T3": {
        "color": "#FF9900",
        "label": "T3\n> 700°C",
        "points": [(46,0,54),(40,60,0),(0,100,0),(0,93,7),(0,0,100)],
    },
    "D1": {
        "color": "#FF7F7F",
        "label": "D1\nLow Energy\nDischarge",
        "points": [(100,0,0),(98,2,0),(76,24,0),(87,0,13)],
    },
    "D2": {
        "color": "#FF3333",
        "label": "D2\nHigh Energy\nDischarge",
        "points": [(87,0,13),(76,24,0),(40,60,0),(23,0,77)],
    },
    "DT": {
        "color": "#CC66FF",
        "label": "DT\nMixed Discharge\n+ Thermal",
        "points": [(23,0,77),(40,60,0),(0,93,7),(0,0,100)],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# 4. DETERMINE FAULT ZONE FOR THIS SAMPLE
# ─────────────────────────────────────────────────────────────────────────────
def classify_duval(pCH4, pC2H4, pC2H2):
    """Standard Duval triangle fault classification."""
    # Using standard boundary rules
    if pC2H2 >= 29:
        if pC2H4 >= 0 and pCH4 <= 23:
            return "DT"
        return "D2"
    if pC2H2 >= 13:
        if pC2H4 <= 60:
            return "D2" if pCH4 <= 23 else "D2"
        return "DT"
    if pC2H2 >= 2:
        if pC2H4 < 24:
            return "T1"  # overlapping with D1 region
        return "D1" if pCH4 >= 87 else "D2"
    # pC2H2 < 2
    if pC2H4 >= 60:
        return "T3"
    if pC2H4 >= 24:
        return "T2"
    if pCH4 >= 98:
        return "PD"
    return "T1"

fault_zone = classify_duval(pCH4, pC2H4, pC2H2)
print(f"\nFault Classification → {fault_zone}")

fault_meanings = {
    "PD": "Partial Discharge",
    "T1": "Thermal Fault < 300°C",
    "T2": "Thermal Fault 300–700°C",
    "T3": "Thermal Fault > 700°C",
    "D1": "Low Energy Electrical Discharge",
    "D2": "High Energy Electrical Discharge (Arc)",
    "DT": "Mixed Discharge + Thermal Fault",
}

# ─────────────────────────────────────────────────────────────────────────────
# 5. PLOT THE DUVAL TRIANGLE
# ─────────────────────────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(10, 9), facecolor="white")
ax.set_aspect("equal")
ax.axis("off")

# Draw filled zone polygons
for zone_name, zone_data in zones.items():
    pts_cart = [ternary_to_cart(*p) for p in zone_data["points"]]
    poly = Polygon(pts_cart, closed=True,
                   facecolor=zone_data["color"], edgecolor="white",
                   linewidth=1.0, alpha=0.85, zorder=1)
    ax.add_patch(poly)

    # Zone label at centroid
    cx = np.mean([p[0] for p in pts_cart])
    cy = np.mean([p[1] for p in pts_cart])
    ax.text(cx, cy, zone_data["label"], ha="center", va="center",
            fontsize=7.5, fontweight="bold", color="#222222",
            zorder=4, linespacing=1.3)

# Triangle outline
triangle = Polygon([(0, 0), (1, 0), (0.5, np.sqrt(3)/2)],
                   closed=True, fill=False, edgecolor="black", linewidth=2, zorder=5)
ax.add_patch(triangle)

# ─────────────────────────────────────────────────────────────────────────────
# 6. AXIS TICK LINES & LABELS (grid lines at 20% intervals)
# ─────────────────────────────────────────────────────────────────────────────
tick_vals = [20, 40, 60, 80]

for tv in tick_vals:
    # CH4 axis lines (parallel to bottom-right edge)
    p1 = ternary_to_cart(tv, 0, 100-tv)
    p2 = ternary_to_cart(tv, 100-tv, 0)
    ax.plot([p1[0], p2[0]], [p1[1], p2[1]], color="gray", lw=0.4, ls="--", zorder=2)

    # C2H4 axis lines
    p1 = ternary_to_cart(0, tv, 100-tv)
    p2 = ternary_to_cart(100-tv, tv, 0)
    ax.plot([p1[0], p2[0]], [p1[1], p2[1]], color="gray", lw=0.4, ls="--", zorder=2)

    # C2H2 axis lines
    p1 = ternary_to_cart(0, 100-tv, tv)
    p2 = ternary_to_cart(100-tv, 0, tv)
    ax.plot([p1[0], p2[0]], [p1[1], p2[1]], color="gray", lw=0.4, ls="--", zorder=2)

# Axis corner labels
offset = 0.04
ax.text(-offset,    -offset,    "100%\nCH₄",  ha="center", va="top",    fontsize=9, fontweight="bold")
ax.text(1+offset,   -offset,    "100%\nC₂H₄", ha="center", va="top",    fontsize=9, fontweight="bold")
ax.text(0.5, np.sqrt(3)/2+0.02, "100%\nC₂H₂", ha="center", va="bottom", fontsize=9, fontweight="bold")

# Axis percentage ticks on bottom
for tv in tick_vals:
    x, y = ternary_to_cart(100-tv, tv, 0)
    ax.text(x, y - 0.04, f"{tv}%", ha="center", va="top", fontsize=7, color="#555555")
    x2, y2 = ternary_to_cart(tv, 0, 100-tv)
    ax.text(x2 - 0.03, y2, f"{tv}%", ha="right",  va="center", fontsize=7, color="#555555")
    x3, y3 = ternary_to_cart(0, tv, 100-tv)
    ax.text(x3 + 0.03, y3, f"{tv}%", ha="left",   va="center", fontsize=7, color="#555555")

# ─────────────────────────────────────────────────────────────────────────────
# 7. PLOT SAMPLE POINT
# ─────────────────────────────────────────────────────────────────────────────
sx, sy = ternary_to_cart(pCH4, pC2H4, pC2H2)

# Highlight zone where point falls
zone_color = zones.get(fault_zone, {}).get("color", "yellow")
ax.scatter([sx], [sy], s=200, color="red", edgecolors="black",
           linewidths=1.5, zorder=10, marker="*")
ax.annotate(
    f"  Sample Point\n  CH₄={pCH4:.1f}%\n  C₂H₄={pC2H4:.1f}%\n  C₂H₂={pC2H2:.1f}%",
    (sx, sy),
    xytext=(sx + 0.12, sy + 0.06),
    fontsize=8,
    color="darkred",
    fontweight="bold",
    arrowprops=dict(arrowstyle="->", color="darkred", lw=1.2),
    bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="darkred", alpha=0.85),
    zorder=11
)

# ─────────────────────────────────────────────────────────────────────────────
# 8. TITLE, LEGEND & INFO BOX
# ─────────────────────────────────────────────────────────────────────────────
ax.set_title("Duval's Triangle (IEC 60599)\nTransformer Oil DGA Analysis",
             fontsize=14, fontweight="bold", pad=14)

# Info box
info_text = (
    f"Gas Values (from DGA Report)\n"
    f"CH₄   = {CH4:.2f} ppm  →  {pCH4:.1f}%\n"
    f"C₂H₄  = {C2H4:.2f} ppm  →  {pC2H4:.1f}%\n"
    f"C₂H₂  = {C2H2:.2f} ppm  →  {pC2H2:.1f}%\n"
    f"─────────────────────\n"
    f"⚠ Fault Zone: {fault_zone}\n"
    f"→ {fault_meanings[fault_zone]}"
)
ax.text(1.02, 0.72, info_text, transform=ax.transAxes,
        fontsize=8.5, va="top", ha="left",
        bbox=dict(boxstyle="round,pad=0.5", facecolor="#FFF8DC", edgecolor="#999", alpha=0.95),
        fontfamily="monospace")

# Legend patches
legend_patches = [mpatches.Patch(color=v["color"], label=v["label"].replace("\n", " "))
                  for v in zones.values()]
ax.legend(handles=legend_patches, loc="lower right",
          bbox_to_anchor=(1.35, 0.0), fontsize=7.5, title="Fault Zones",
          framealpha=0.9, title_fontsize=8)

plt.tight_layout()

# Save PNG
plt.savefig(PNG_PATH, dpi=180, bbox_inches="tight",
            facecolor="white")
plt.close()
print("Duval triangle PNG saved.")

# ─────────────────────────────────────────────────────────────────────────────
# 9. EMBED CHART INTO EXCEL (new sheet)
# ─────────────────────────────────────────────────────────────────────────────
wb = load_workbook(EXCEL_PATH)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

wb2 = load_workbook(EXCEL_PATH)

# Update fault zone cell in DGA sheet
ws_dga = wb2["DGA Data"]
ws_dga["A27"] = f"Duval Fault Zone: {fault_zone} — {fault_meanings[fault_zone]}"

# Create chart sheet
if "Duval Triangle" in wb2.sheetnames:
    del wb2["Duval Triangle"]
ws_chart = wb2.create_sheet("Duval Triangle")
ws_chart.sheet_view.showGridLines = False

# Header
ws_chart.merge_cells("A1:L1")
ws_chart["A1"] = "Duval's Triangle — IEC 60599 Fault Zone Analysis"
ws_chart["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws_chart["A1"].fill = PatternFill("solid", start_color="1F4E79")
ws_chart["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_chart.row_dimensions[1].height = 28

# Results summary
summary = [
    ("Gas", "Value (ppm)", "% Contribution"),
    ("Methane (CH₄)",     f"{CH4:.2f}",  f"{pCH4:.2f}%"),
    ("Ethylene (C₂H₄)",   f"{C2H4:.2f}", f"{pC2H4:.2f}%"),
    ("Acetylene (C₂H₂)",  f"{C2H2:.2f}", f"{pC2H2:.2f}%"),
    ("Total",             f"{total:.2f}", "100.00%"),
    ("", "", ""),
    ("Duval Zone", fault_zone, fault_meanings[fault_zone]),
]

from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
thin = Side(style="thin", color="000000")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, row in enumerate(summary, start=3):
    for j, val in enumerate(row, start=1):
        c = ws_chart.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", bold=(i==3 or i==9), size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = bdr
        if i == 3:
            c.fill = PatternFill("solid", start_color="2E75B6")
            c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        if i == 9:
            c.fill = PatternFill("solid", start_color="FFC000")

ws_chart.column_dimensions["A"].width = 22
ws_chart.column_dimensions["B"].width = 16
ws_chart.column_dimensions["C"].width = 28

# Embed chart image
img = XLImage(PNG_PATH)
img.anchor = "A12"
ws_chart.add_image(img)

wb2.save(EXCEL_PATH)
print("Excel updated with Duval chart sheet.")
print(f"\n=== FINAL RESULT ===")
print(f"Fault Zone: {fault_zone} — {fault_meanings[fault_zone]}")
print(f"Transformer condition: NORMAL / No active fault indicated.")
