"""
Full Dissolved Gas Analysis (DGA) — Transformer Oil
=====================================================
Methods covered:
  1. Duval's Triangle (IEC 60599)
  2. Rogers Ratio Method (IEC 60599)
  3. Doernenburg Ratio Method
  4. Key Gas Method (IEEE C57.104)
  5. IEC 60599 Ratio Method (Revised)
  6. TDCG / IEEE C57.104 Condition Assessment
  7. CO2/CO Ratio (Cellulose Degradation)
  8. Individual Gas Limit Check (IS 9434)

All results are written to a multi-sheet Excel report and
all charts are saved as PNG files in the same folder.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(BASE_DIR, "Full_DGA_Report.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# ① GAS VALUES  ←  EDIT THESE FOR YOUR TRANSFORMER
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# READ GAS VALUES FROM EXCEL
# ─────────────────────────────────────────────────────────────────────────────

INPUT_EXCEL = os.path.join(BASE_DIR, "DGA_Input.xlsx")

df = pd.read_excel(INPUT_EXCEL)

# Read first row from Excel

H2   = float(df.loc[0, "H2"])
O2   = float(df.loc[0, "O2"])
N2   = float(df.loc[0, "N2"])
CO   = float(df.loc[0, "CO"])
CH4  = float(df.loc[0, "CH4"])
CO2  = float(df.loc[0, "CO2"])
C2H4 = float(df.loc[0, "C2H4"])
C2H6 = float(df.loc[0, "C2H6"])
C2H2 = float(df.loc[0, "C2H2"])
C3H6 = float(df.loc[0, "C3H6"])
C3H8 = float(df.loc[0, "C3H8"])
# Derived totals
TDCG  = H2 + CO + CH4 + C2H4 + C2H6 + C2H2   # Total Dissolved Combustible Gas
TGC_pct = 5.0                                   # Total Gas Content % v/v (from report)

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, cell, val, bg="1F4E79", fg="FFFFFF", sz=11, bold=True, wrap=False):
    c = ws[cell]
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = thin_border()

def cell_write(ws, row, col, val, bold=False, color="000000", bg=None,
               align="center", fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

STATUS_COLOR = {"Acceptable": "E2EFDA", "Warning": "FFEB9C",
                "Alarm": "FFC7CE", "Not Specified": "F2F2F2",
                "Normal": "E2EFDA", "Abnormal": "FFC7CE"}

# ─────────────────────────────────────────────────────────────────────────────
# ② INDIVIDUAL GAS LIMITS  (IS 9434 / IEEE C57.104)
# ─────────────────────────────────────────────────────────────────────────────
GAS_LIMITS = {
    # (value, limit_val, limit_str, unit)
    "H₂  (Hydrogen)":         (H2,   50,   "50 Max",   "ppm"),
    "O₂  (Oxygen)":           (O2,   None, "NS",       "ppm"),
    "N₂  (Nitrogen)":         (N2,   None, "NS",       "ppm"),
    "CO  (Carbon Monoxide)":  (CO,   400,  "400 Max",  "ppm"),
    "CH₄ (Methane)":          (CH4,  30,   "30 Max",   "ppm"),
    "CO₂ (Carbon Dioxide)":   (CO2,  3800, "3800 Max", "ppm"),
    "C₂H₄ (Ethylene)":        (C2H4, 60,   "60 Max",   "ppm"),
    "C₂H₆ (Ethane)":          (C2H6, 20,   "20 Max",   "ppm"),
    "C₂H₂ (Acetylene)":       (C2H2, 2,    "2 Max",    "ppm"),
    "C₃H₆ (Propylene)":       (C3H6, None, "NS",       "ppm"),
    "C₃H₈ (Propane)":         (C3H8, None, "NS",       "ppm"),
    "TDCG":                   (TDCG, None, "NS",       "ppm"),
}

def gas_status(val, limit):
    if limit is None:
        return "Not Specified"
    return "Acceptable" if val <= limit else "Alarm"

# ─────────────────────────────────────────────────────────────────────────────
# ③ TDCG / IEEE C57.104 CONDITION LEVELS
# ─────────────────────────────────────────────────────────────────────────────
def tdcg_condition(tdcg_val):
    if tdcg_val <= 720:
        return 1, "Normal Operation", "No action required.", "E2EFDA"
    elif tdcg_val <= 1920:
        return 2, "Caution", "Resample & re-test in 1 month.", "FFEB9C"
    elif tdcg_val <= 4630:
        return 3, "Warning", "Resample in 1 week. Plan inspection.", "FFC7CE"
    else:
        return 4, "Critical", "Immediate action / take offline.", "FF0000"

# ─────────────────────────────────────────────────────────────────────────────
# ④ CO₂/CO RATIO — Cellulose / Paper Degradation
# ─────────────────────────────────────────────────────────────────────────────
def co2_co_ratio(co2_val, co_val):
    if co_val == 0:
        return None, "Indeterminate (CO = 0)"
    r = co2_val / co_val
    if r < 3:
        return r, "Possible cellulose fault (overheating of paper insulation)"
    elif r <= 10:
        return r, "Normal cellulose aging — no active paper fault"
    else:
        return r, "Possible CO₂ from non-fault source OR extensive paper degradation"

# ─────────────────────────────────────────────────────────────────────────────
# ⑤ ROGERS RATIO METHOD (IEC 60599)
# ─────────────────────────────────────────────────────────────────────────────
def rogers_ratio(h2, ch4, c2h2, c2h4, c2h6):
    """Returns (fault_code, fault_description, ratios_dict)"""
    r1 = c2h2 / c2h4 if c2h4 > 0 else 0    # C2H2/C2H4
    r2 = ch4  / h2   if h2   > 0 else 0    # CH4/H2
    r5 = c2h4 / c2h6 if c2h6 > 0 else 0    # C2H4/C2H6

    def code(val, thresholds):
        """Map ratio to Rogers code (0/1/2)."""
        lo, hi = thresholds
        if val < lo:   return 0
        elif val < hi: return 1
        else:          return 2

    c1 = code(r1, (0.1, 3.0))
    c2 = code(r2, (0.1, 1.0))
    c5 = code(r5, (1.0, 3.0))

    cases = {
        (0,0,0): "Normal Aging",
        (0,1,0): "Partial Discharge (low energy density)",
        (1,1,0): "Partial Discharge (high energy density / tracking)",
        (1,0,1): "Low Energy Discharge (arcing, corona)",
        (1,0,2): "High Energy Discharge (arcing)",
        (0,0,1): "Thermal Fault < 150°C (slight overheating)",
        (0,0,2): "Thermal Fault 150–200°C",
        (0,1,1): "Thermal Fault 200–300°C",
        (0,1,2): "Thermal Fault > 300°C (conductor overheating)",
    }
    key = (c1, c2, c5)
    fault = cases.get(key, f"Unclassified (codes: C2H2/C2H4={c1}, CH4/H2={c2}, C2H4/C2H6={c5})")
    return fault, {"C₂H₂/C₂H₄": r1, "CH₄/H₂": r2, "C₂H₄/C₂H₆": r5}

# ─────────────────────────────────────────────────────────────────────────────
# ⑥ DOERNENBURG RATIO METHOD
# ─────────────────────────────────────────────────────────────────────────────
def doernenburg(h2, ch4, c2h2, c2h4, c2h6, co):
    """
    Requires minimum gas concentrations to be valid.
    Returns (fault, ratios, valid)
    """
    min_vals = {"H2": 100, "CH4": 120, "C2H2": 35, "C2H4": 50}
    valid = (h2 >= min_vals["H2"] or ch4 >= min_vals["CH4"] or
             c2h2 >= min_vals["C2H2"] or c2h4 >= min_vals["C2H4"])

    r1 = ch4  / h2   if h2   > 0 else 0   # CH4/H2
    r2 = c2h2 / c2h4 if c2h4 > 0 else 0   # C2H2/C2H4
    r3 = c2h2 / ch4  if ch4  > 0 else 0   # C2H2/CH4
    r4 = c2h6 / c2h2 if c2h2 > 0 else 0   # C2H6/C2H2 (inverse for check)

    if not valid:
        fault = "Insufficient gas levels — Doernenburg method not applicable"
    elif r1 > 1.0 and r2 < 0.75 and r3 < 0.3:
        fault = "Thermal Decomposition"
    elif r1 < 0.1 and r2 < 0.75 and r3 < 0.3:
        fault = "Partial Discharge (corona)"
    elif r2 >= 0.75 and r3 >= 0.3:
        fault = "Arcing (high energy discharge)"
    else:
        fault = "Normal / No significant fault"

    return fault, {"CH₄/H₂": r1, "C₂H₂/C₂H₄": r2, "C₂H₂/CH₄": r3}, valid

# ─────────────────────────────────────────────────────────────────────────────
# ⑦ IEC 60599 REVISED RATIO METHOD
# ─────────────────────────────────────────────────────────────────────────────
def iec_60599(h2, ch4, c2h2, c2h4, c2h6):
    r1 = c2h2 / c2h4 if c2h4 > 0 else 0
    r2 = ch4  / h2   if h2   > 0 else 0

    if r1 < 0.1:
        if r2 < 0.1:
            fault = "Partial Discharge"
        elif r2 < 1.0:
            fault = "Thermal Fault (stray gassing / low temp)"
        else:
            fault = "Thermal Fault (high temperature > 300°C)"
    elif r1 < 3.0:
        if r2 < 0.1:
            fault = "Discharge of Low Energy (sparking)"
        else:
            fault = "Discharge of High Energy (arcing)"
    else:
        fault = "Discharge of High Energy (severe arcing)"

    return fault, {"C₂H₂/C₂H₄": r1, "CH₄/H₂": r2}

# ─────────────────────────────────────────────────────────────────────────────
# ⑧ KEY GAS METHOD (IEEE C57.104 / Transformer Aging Guide)
# ─────────────────────────────────────────────────────────────────────────────
def key_gas_method(h2, ch4, c2h2, c2h4, c2h6, co):
    """
    Identifies dominant fault type from the single highest key gas.
    """
    gases = {
        "H₂":   h2,
        "CH₄":  ch4,
        "C₂H₄": c2h4,
        "C₂H₂": c2h2,
        "C₂H₆": c2h6,
        "CO":   co,
    }
    key = max(gases, key=gases.get)
    val = gases[key]

    interpretations = {
        "H₂":   "Corona / Partial Discharge (low-energy sparking in oil)",
        "CH₄":  "Low-temperature thermal fault (< 300°C in oil)",
        "C₂H₄": "High-temperature thermal fault in oil (> 300°C)",
        "C₂H₂": "Electrical arcing / high-energy discharge",
        "C₂H₆": "Low-to-moderate thermal fault in oil (< 200°C)",
        "CO":   "Thermal fault involving cellulose (paper/pressboard overheating)",
    }
    return key, val, interpretations[key], gases


DUVAL_ZONES = {
    "PD": {"color":"#B3D9FF","label":"PD\nPartial\nDischarge",
           "points":[(98,0,2),(100,0,0),(98,2,0)]},
    "T1": {"color":"#FFFF99","label":"T1\n<300°C",
           "points":[(98,0,2),(98,2,0),(76,24,0),(77,0,23)]},
    "T2": {"color":"#FFD966","label":"T2\n300–700°C",
           "points":[(77,0,23),(76,24,0),(40,60,0),(46,0,54)]},
    "T3": {"color":"#FF9900","label":"T3\n>700°C",
           "points":[(46,0,54),(40,60,0),(0,100,0),(0,93,7),(0,0,100)]},
    "D1": {"color":"#FF7F7F","label":"D1\nLow Energy\nDischarge",
           "points":[(100,0,0),(98,2,0),(76,24,0),(87,0,13)]},
    "D2": {"color":"#FF3333","label":"D2\nHigh Energy\nDischarge",
           "points":[(87,0,13),(76,24,0),(40,60,0),(23,0,77)]},
    "DT": {"color":"#CC66FF","label":"DT\nMixed\nDisch+Therm",
           "points":[(23,0,77),(40,60,0),(0,93,7),(0,0,100)]},
}

DUVAL_MEANINGS = {

    "NORMAL": "Insufficient gas for reliable Duval analysis",

    "PD": "Partial Discharge",
    "T1": "Thermal Fault < 300°C",
    "T2": "Thermal Fault 300–700°C",
    "T3": "Thermal Fault > 700°C",
    "D1": "Low Energy Discharge",
    "D2": "High Energy Arcing",
    "DT": "Mixed Thermal + Electrical Fault"
}

def duval_analysis(ch4_v, c2h4_v, c2h2_v):

    total = ch4_v + c2h4_v + c2h2_v

    if total == 0:
        pCH4 = 33.3
        pC2H4 = 33.3
        pC2H2 = 33.3

    else:
        pCH4  = ch4_v  / total * 100
        pC2H4 = c2h4_v / total * 100
        pC2H2 = c2h2_v / total * 100

    fault_zone = classify_duval(
        pCH4,
        pC2H4,
        pC2H2,
        total
    )

    return fault_zone, pCH4, pC2H4, pC2H2
    
def classify_duval(pCH4, pC2H4, pC2H2, total_ppm):

    # Ignore Duval when gas quantity is too low
    if total_ppm < 50:
        return "NORMAL"

    if pC2H2 >= 29:
        return "DT" if pCH4 <= 23 else "D2"

    if pC2H2 >= 13:
        return "D2"

    if pC2H2 >= 2:
        if pC2H4 < 24:
            return "T1"

        return "D1" if pCH4 >= 87 else "D2"

    if pC2H4 >= 60:
        return "T3"

    if pC2H4 >= 24:
        return "T2"

    if pCH4 >= 98:
        return "PD"

    return "T1"


# ═════════════════════════════════════════════════════════════════════════════
# RUN ALL ANALYSES
# ═════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("  FULL DGA ANALYSIS — TRANSFORMER OIL")
print("=" * 60)

# Duval
duval_zone, pCH4, pC2H4, pC2H2 = duval_analysis(CH4, C2H4, C2H2)
print(f"\n[Duval Triangle]  Zone: {duval_zone} — {DUVAL_MEANINGS[duval_zone]}")

# Rogers
rogers_fault, rogers_ratios = rogers_ratio(H2, CH4, C2H2, C2H4, C2H6)
print(f"[Rogers Ratio]    {rogers_fault}")

# Doernenburg
doern_fault, doern_ratios, doern_valid = doernenburg(H2, CH4, C2H2, C2H4, C2H6, CO)
print(f"[Doernenburg]     {doern_fault}")

# IEC 60599
iec_fault, iec_ratios = iec_60599(H2, CH4, C2H2, C2H4, C2H6)
print(f"[IEC 60599]       {iec_fault}")

# Key Gas
key_gas, key_val, key_interp, all_key_gases = key_gas_method(H2, CH4, C2H2, C2H4, C2H6, CO)
print(f"[Key Gas]         Dominant: {key_gas} = {key_val:.2f} ppm → {key_interp}")

# TDCG
tdcg_level, tdcg_status, tdcg_action, tdcg_color = tdcg_condition(TDCG)
print(f"[TDCG]            {TDCG:.2f} ppm → Level {tdcg_level}: {tdcg_status}")

# CO2/CO
co2co_ratio, co2co_interp = co2_co_ratio(CO2, CO)
if co2co_ratio is not None:
    print(f"[CO₂/CO Ratio]    {co2co_ratio:.2f} → {co2co_interp}")
else:
    print(f"[CO₂/CO Ratio]    N/A → {co2co_interp}")

# Severity scores for radar (0=normal, 1=critical)
severity_map = {
    "Normal Aging":0.05, "Partial Discharge (low energy density)":0.3,
    "Partial Discharge (high energy density / tracking)":0.5,
    "Low Energy Discharge (arcing, corona)":0.6,
    "High Energy Discharge (arcing)":0.9,
    "Thermal Fault < 150°C (slight overheating)":0.2,
    "Thermal Fault 150–200°C":0.35, "Thermal Fault 200–300°C":0.45,
    "Thermal Fault > 300°C (conductor overheating)":0.65,
}
duval_sev  = {"PD":0.4,"T1":0.2,"T2":0.45,"T3":0.7,"D1":0.55,"D2":0.85,"DT":0.75}
tdcg_sev   = {1:0.05, 2:0.35, 3:0.65, 4:1.0}
radar_scores = [
    duval_sev.get(duval_zone, 0.1),
    severity_map.get(rogers_fault, 0.1),
    0.1 if not doern_valid else (0.7 if "Arcing" in doern_fault else 0.1),
    0.1 if "Normal" in iec_fault or "stray" in iec_fault else 0.5,
    0.1 if key_gas in ("C₂H₆","CH₄") else (0.8 if key_gas=="C₂H₂" else 0.3),
    tdcg_sev.get(tdcg_level, 0.1),
]
radar_labels = ["Duval\nTriangle","Rogers\nRatio","Doernenburg",
                "IEC 60599","Key Gas\nMethod","TDCG\nIEEE C57.104"]

# ═════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL WORKBOOK
# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─── Sheet 1: GAS DATA ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "1_Gas Data"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:G1")
hdr(ws1,"A1","DGA REPORT — Full Gas Concentration Data (IS 9434 / IEEE C57.104)",
    bg="1F4E79", sz=13)
ws1.row_dimensions[1].height = 30

col_heads = ["Parameter","Symbol","Unit","Measured Value","Limit","Status","Remark"]
for j, h in enumerate(col_heads, 1):
    hdr(ws1, f"{get_column_letter(j)}3", h, bg="2E75B6", sz=10)
ws1.row_dimensions[3].height = 22

gas_rows = [
    ("Total Gas Content",       "TGC",    "% v/v",  TGC_pct, None,  "NS",       ""),
    ("Hydrogen",                "H₂",     "ppm",    H2,      50,    "50 Max",   "Key fault gas"),
    ("Oxygen",                  "O₂",     "ppm",    O2,      None,  "NS",       "Monitor O2/N2 ratio"),
    ("Nitrogen",                "N₂",     "ppm",    N2,      None,  "NS",       ""),
    ("Carbon Monoxide",         "CO",     "ppm",    CO,      400,   "400 Max",  "Paper degradation indicator"),
    ("Methane",                 "CH₄",    "ppm",    CH4,     30,    "30 Max",   "Key fault gas"),
    ("Carbon Dioxide",          "CO₂",    "ppm",    CO2,     3800,  "3800 Max", "Paper degradation indicator"),
    ("Ethylene",                "C₂H₄",   "ppm",    C2H4,    60,    "60 Max",   "Key fault gas"),
    ("Ethane",                  "C₂H₆",   "ppm",    C2H6,    20,    "20 Max",   "Key fault gas"),
    ("Acetylene",               "C₂H₂",   "ppm",    C2H2,    2,     "2 Max",    "Key fault gas — arcing"),
    ("Propylene",               "C₃H₆",   "ppm",    C3H6,    None,  "NS",       ""),
    ("Propane",                 "C₃H₈",   "ppm",    C3H8,    None,  "NS",       ""),
    ("Total Dissolved Comb Gas","TDCG",   "ppm",    TDCG,    720,   "720 Max",  "IEEE Level 1 threshold"),
    ("CO₂/CO Ratio",           "CO₂/CO", "—",
     round(CO2/CO,3) if CO > 0 else 0, None, "3–10 Normal", "Cellulose degradation"),
    ("O₂/N₂ Ratio",            "O₂/N₂",  "—",
     round(O2/N2,4) if N2 > 0 else 0, None, "> 0.3 Normal", "Breathing indicator"),
]

for i, row_data in enumerate(gas_rows, start=4):
    name, sym, unit, val, lim, lim_str, remark = row_data
    status = gas_status(val, lim) if lim is not None else "Not Specified"
    bg = STATUS_COLOR.get(status, "FFFFFF")
    cell_write(ws1, i, 1, name,    align="left",   bg=None)
    cell_write(ws1, i, 2, sym,     bg=None)
    cell_write(ws1, i, 3, unit,    bg=None)
    cell_write(ws1, i, 4, val,     fmt="0.000",    bg=None, color="0000FF")
    cell_write(ws1, i, 5, lim_str, bg=None)
    cell_write(ws1, i, 6, status,  bg=bg, bold=True)
    cell_write(ws1, i, 7, remark,  align="left",   bg=None)

col_widths1 = [28, 10, 8, 18, 14, 16, 32]
for j, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(j)].width = w

# ─── Sheet 2: DUVAL TRIANGLE ──────────────────────────────────────────────
ws2 = wb.create_sheet("2_Duval Triangle")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:E1")
hdr(ws2,"A1","DUVAL'S TRIANGLE — IEC 60599 Fault Zone Analysis", bg="1F4E79", sz=13)
ws2.row_dimensions[1].height = 30

summary_data = [
    ("Gas","Value (ppm)","% of Sum","",""),
    ("Methane CH₄",  f"{CH4:.3f}",  f"{pCH4:.2f}%", "",""),
    ("Ethylene C₂H₄",f"{C2H4:.3f}", f"{pC2H4:.2f}%","",""),
    ("Acetylene C₂H₂",f"{C2H2:.3f}",f"{pC2H2:.2f}%","",""),
    ("Total",        f"{CH4+C2H4+C2H2:.3f}","100.00%","",""),
    ("","","","",""),
    ("Fault Zone", duval_zone, DUVAL_MEANINGS[duval_zone],"",""),
]
for i, row in enumerate(summary_data, start=3):
    for j, v in enumerate(row[:3], start=1):
        bg = "FFC000" if i == 9 else ("2E75B6" if i == 3 else None)
        fg = "FFFFFF" if i == 3 else "000000"
        cell_write(ws2, i, j, v, bold=(i==3 or i==9), bg=bg, color=fg)

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 32

# ─── Sheet 3: RATIO METHODS ───────────────────────────────────────────────
ws3 = wb.create_sheet("3_Ratio Methods")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
hdr(ws3,"A1","RATIO METHODS — Rogers / Doernenburg / IEC 60599", bg="1F4E79", sz=13)
ws3.row_dimensions[1].height = 30

# Section headers
def section_hdr(ws, row, text, bg="375623"):
    ws.merge_cells(f"A{row}:F{row}")
    hdr(ws, f"A{row}", text, bg=bg, sz=11)
    ws.row_dimensions[row].height = 20

# Rogers
section_hdr(ws3, 3, "ROGERS RATIO METHOD (IEC 60599)", "17375E")
hdr(ws3,"A4","Ratio",    bg="2E75B6", sz=10)
hdr(ws3,"B4","Value",    bg="2E75B6", sz=10)
hdr(ws3,"C4","Meaning",  bg="2E75B6", sz=10)
hdr(ws3,"D4","Code Range",bg="2E75B6",sz=10)

rogers_info = [
    ("C₂H₂/C₂H₄", rogers_ratios["C₂H₂/C₂H₄"], "Discharge indicator", "0.1–3.0"),
    ("CH₄/H₂",     rogers_ratios["CH₄/H₂"],     "Thermal vs PD",       "0.1–1.0"),
    ("C₂H₄/C₂H₆",  rogers_ratios["C₂H₄/C₂H₆"],  "Thermal level",       "1.0–3.0"),
]
for i, (ratio, val, meaning, rng) in enumerate(rogers_info, start=5):
    cell_write(ws3, i, 1, ratio,   align="left")
    cell_write(ws3, i, 2, round(val,4), fmt="0.0000", color="0000FF")
    cell_write(ws3, i, 3, meaning, align="left")
    cell_write(ws3, i, 4, rng)

ws3.merge_cells("A8:F8")
c = ws3["A8"]
c.value = f"Rogers Diagnosis: {rogers_fault}"
c.font = Font(name="Arial", bold=True, size=11, color="000000")
c.fill = PatternFill("solid", start_color="FFD966")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws3.row_dimensions[8].height = 22

# Doernenburg
section_hdr(ws3, 10, "DOERNENBURG RATIO METHOD", "17375E")
hdr(ws3,"A11","Ratio",   bg="2E75B6", sz=10)
hdr(ws3,"B11","Value",   bg="2E75B6", sz=10)
hdr(ws3,"C11","Threshold",bg="2E75B6",sz=10)
hdr(ws3,"D11","Valid?",  bg="2E75B6", sz=10)

doern_info = [
    ("CH₄/H₂",    doern_ratios["CH₄/H₂"],    "> 1.0"),
    ("C₂H₂/C₂H₄", doern_ratios["C₂H₂/C₂H₄"], "≥ 0.75"),
    ("C₂H₂/CH₄",  doern_ratios["C₂H₂/CH₄"],  "≥ 0.3"),
]
for i, (ratio, val, thr) in enumerate(doern_info, start=12):
    cell_write(ws3, i, 1, ratio,  align="left")
    cell_write(ws3, i, 2, round(val,4), fmt="0.0000", color="0000FF")
    cell_write(ws3, i, 3, thr)
    cell_write(ws3, i, 4, "Yes" if doern_valid else "No (insufficient gas levels)")

ws3.merge_cells("A15:F15")
c = ws3["A15"]
c.value = f"Doernenburg Diagnosis: {doern_fault}"
c.font = Font(name="Arial", bold=True, size=11)
c.fill = PatternFill("solid", start_color="FFD966")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws3.row_dimensions[15].height = 22

# IEC 60599
section_hdr(ws3, 17, "IEC 60599 REVISED RATIO METHOD", "17375E")
hdr(ws3,"A18","Ratio",   bg="2E75B6", sz=10)
hdr(ws3,"B18","Value",   bg="2E75B6", sz=10)
hdr(ws3,"C18","Zone",    bg="2E75B6", sz=10)

iec_info = [
    ("C₂H₂/C₂H₄", iec_ratios["C₂H₂/C₂H₄"], "< 0.1 | 0.1–3 | > 3"),
    ("CH₄/H₂",     iec_ratios["CH₄/H₂"],     "< 0.1 | 0.1–1 | > 1"),
]
for i, (ratio, val, zone) in enumerate(iec_info, start=19):
    cell_write(ws3, i, 1, ratio,  align="left")
    cell_write(ws3, i, 2, round(val,4), fmt="0.0000", color="0000FF")
    cell_write(ws3, i, 3, zone)

ws3.merge_cells("A21:F21")
c = ws3["A21"]
c.value = f"IEC 60599 Diagnosis: {iec_fault}"
c.font = Font(name="Arial", bold=True, size=11)
c.fill = PatternFill("solid", start_color="FFD966")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws3.row_dimensions[21].height = 22

for j, w in enumerate([20,14,20,18,14,14],1):
    ws3.column_dimensions[get_column_letter(j)].width = w

# ─── Sheet 4: KEY GAS + TDCG ──────────────────────────────────────────────
ws4 = wb.create_sheet("4_KeyGas_TDCG")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:E1")
hdr(ws4,"A1","KEY GAS METHOD & TDCG ASSESSMENT (IEEE C57.104)", bg="1F4E79", sz=13)
ws4.row_dimensions[1].height = 30

# Key Gas table
section_hdr(ws4, 3, "KEY GAS METHOD (IEEE C57.104 / Transformer Guide)", "17375E")
hdr(ws4,"A4","Gas",    bg="2E75B6", sz=10)
hdr(ws4,"B4","Value (ppm)", bg="2E75B6", sz=10)
hdr(ws4,"C4","Fault Indication", bg="2E75B6", sz=10)
hdr(ws4,"D4","Dominant?", bg="2E75B6", sz=10)

key_gas_table = [
    ("H₂",   H2,   "Corona / Partial Discharge"),
    ("CH₄",  CH4,  "Low-temp thermal (< 300°C)"),
    ("C₂H₄", C2H4, "High-temp thermal in oil (> 300°C)"),
    ("C₂H₂", C2H2, "Electrical arcing"),
    ("C₂H₆", C2H6, "Low-moderate thermal (< 200°C)"),
    ("CO",   CO,   "Paper/cellulose overheating"),
]
for i, (g, v, f) in enumerate(key_gas_table, start=5):
    is_dominant = (g == key_gas)
    bg = "FFC000" if is_dominant else None
    cell_write(ws4, i, 1, g, bold=is_dominant, bg=bg)
    cell_write(ws4, i, 2, v, fmt="0.000", color="0000FF", bg=bg)
    cell_write(ws4, i, 3, f, align="left", bg=bg)
    cell_write(ws4, i, 4, "★ YES" if is_dominant else "—", bold=is_dominant, bg=bg)

ws4.merge_cells("A12:D12")
c = ws4["A12"]
c.value = f"Key Gas Diagnosis:  {key_gas} dominant → {key_interp}"
c.font = Font(name="Arial", bold=True, size=11)
c.fill = PatternFill("solid", start_color="FFD966")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws4.row_dimensions[12].height = 22

# TDCG Section
section_hdr(ws4, 14, "TDCG ASSESSMENT — IEEE C57.104", "17375E")
hdr(ws4,"A15","Level",  bg="2E75B6", sz=10)
hdr(ws4,"B15","TDCG Range (ppm)", bg="2E75B6", sz=10)
hdr(ws4,"C15","Condition",  bg="2E75B6", sz=10)
hdr(ws4,"D15","Recommended Action", bg="2E75B6", sz=10)

tdcg_levels = [
    (1, "≤ 720",        "Normal",   "No action required"),
    (2, "721 – 1920",   "Caution",  "Resample in 1 month"),
    (3, "1921 – 4630",  "Warning",  "Resample in 1 week; plan inspection"),
    (4, "> 4630",       "Critical", "Immediate action / consider taking offline"),
]
level_colors = {"Normal":"E2EFDA","Caution":"FFEB9C","Warning":"FFC7CE","Critical":"FF0000"}
for i, (lv, rng, cond, action) in enumerate(tdcg_levels, start=16):
    is_current = (lv == tdcg_level)
    bg = level_colors[cond] if is_current else None
    cell_write(ws4, i, 1, f"Level {lv}", bold=is_current, bg=bg)
    cell_write(ws4, i, 2, rng,   bold=is_current, bg=bg)
    cell_write(ws4, i, 3, cond,  bold=is_current, bg=bg)
    cell_write(ws4, i, 4, action,bold=is_current, bg=bg, align="left")

ws4.merge_cells("A21:D21")
c = ws4["A21"]
c.value = (f"TDCG = {TDCG:.2f} ppm  →  Level {tdcg_level}: {tdcg_status}  |  "
           f"Action: {tdcg_action}")
c.font = Font(name="Arial", bold=True, size=11)
c.fill = PatternFill("solid", start_color=tdcg_color)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws4.row_dimensions[21].height = 22

# CO2/CO
section_hdr(ws4, 23, "CO₂/CO RATIO — Cellulose / Paper Insulation Assessment", "17375E")
cell_write(ws4, 24, 1, "CO₂",  bold=True, bg="D9E1F2")
cell_write(ws4, 24, 2, CO2,    fmt="0.00", color="0000FF")
cell_write(ws4, 25, 1, "CO",   bold=True, bg="D9E1F2")
cell_write(ws4, 25, 2, CO,     fmt="0.00", color="0000FF")
cell_write(ws4, 26, 1, "CO₂/CO Ratio", bold=True, bg="FFC000")
cell_write(ws4, 26, 2, round(co2co_ratio,3) if co2co_ratio else "N/A",
           bold=True, bg="FFC000")
ws4.merge_cells("A27:D27")
c = ws4["A27"]
c.value = f"Interpretation: {co2co_interp}"
c.font = Font(name="Arial", bold=True, size=10)
c.fill = PatternFill("solid", start_color="FFD966")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

for j, w in enumerate([20,18,22,36],1):
    ws4.column_dimensions[get_column_letter(j)].width = w

# ─── Sheet 5: SUMMARY & CHARTS ────────────────────────────────────────────
ws5 = wb.create_sheet("5_Summary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:F1")
hdr(ws5,"A1","DGA ANALYSIS — CONSOLIDATED SUMMARY", bg="1F4E79", sz=14)
ws5.row_dimensions[1].height = 32

hdr(ws5,"A3","Method",     bg="2E75B6", sz=10)
hdr(ws5,"B3","Result",     bg="2E75B6", sz=10)
hdr(ws5,"C3","Severity",   bg="2E75B6", sz=10)
hdr(ws5,"D3","Remarks",    bg="2E75B6", sz=10)
ws5.row_dimensions[3].height = 20

sev_label = lambda s: "Normal" if s<0.2 else ("Low" if s<0.4 else ("Moderate" if s<0.6 else ("High" if s<0.8 else "Critical")))
sev_bg    = lambda s: "E2EFDA" if s<0.2 else ("FFEB9C" if s<0.4 else ("FFD966" if s<0.6 else ("FFC7CE" if s<0.8 else "FF0000")))

summary_rows = [
    ("Duval Triangle",       f"{duval_zone} — {DUVAL_MEANINGS[duval_zone]}",  radar_scores[0], "IEC 60599"),
    ("Rogers Ratio",         rogers_fault,                                     radar_scores[1], "IEC 60599"),
    ("Doernenburg",          doern_fault,                                      radar_scores[2], "Min gas threshold may not be met"),
    ("IEC 60599 Revised",    iec_fault,                                        radar_scores[3], "IEC 60599"),
    ("Key Gas",              f"{key_gas} dominant — {key_interp}",            radar_scores[4], "IEEE C57.104"),
    ("TDCG",                 f"Level {tdcg_level}: {tdcg_status} ({TDCG:.2f} ppm)",radar_scores[5],"IEEE C57.104"),
    ("CO₂/CO Ratio",         co2co_interp,                                     0.1,             f"Ratio = {co2co_ratio:.2f}" if co2co_ratio else "CO=0"),
]
for i, (method, result, sev, remark) in enumerate(summary_rows, start=4):
    cell_write(ws5, i, 1, method, bold=True,  align="left")
    cell_write(ws5, i, 2, result, align="left", wrap=True)
    cell_write(ws5, i, 3, sev_label(sev), bold=True, bg=sev_bg(sev))
    cell_write(ws5, i, 4, remark, align="left")
    ws5.row_dimensions[i].height = 20

# Overall
overall_sev = max(radar_scores)
ws5.merge_cells("A12:D12")
c = ws5["A12"]
c.value = (f"OVERALL ASSESSMENT:  {sev_label(overall_sev).upper()}  "
           f"(Severity Index = {overall_sev:.2f})")
c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", start_color="1F4E79")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()
ws5.row_dimensions[12].height = 26

for j, w in enumerate([22,48,14,30],1):
    ws5.column_dimensions[get_column_letter(j)].width = w

# ─── Save workbook ────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)

print(f"\nExcel saved → {EXCEL_PATH}")
print("\n" + "="*60)
print("  OVERALL ASSESSMENT")
print("="*60)
print(f"  Severity: {sev_label(overall_sev).upper()} (index={overall_sev:.2f})")
print(f"  Duval   : {duval_zone} — {DUVAL_MEANINGS[duval_zone]}")
print(f"  Rogers  : {rogers_fault}")
print(f"  IEC     : {iec_fault}")
print(f"  TDCG    : Level {tdcg_level} — {tdcg_status}")
print(f"  CO₂/CO  : {co2co_ratio:.2f} → {co2co_interp}" if co2co_ratio else "  CO₂/CO  : N/A")
print("="*60)
print("\nDone. Open Full_DGA_Report.xlsx in your folder.")