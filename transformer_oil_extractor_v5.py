#!/usr/bin/env python3
"""
Transformer Oil Test Report Extractor
======================================
Reads TRU-FIL PDF oil test reports and appends extracted data into an Excel file.
Keeps looping until you type 'exit'.

Usage:
    python3 transformer_oil_extractor.py

Dependencies (auto-installed if missing):
    pip install pdfplumber openpyxl
"""

import os, re, sys
from tkinter import Tk
from tkinter.filedialog import askopenfilenames


EXCEL_FILE = "transformer_oil_data.xlsx"

HEADERS = [
    "Owner",
    "Installation Location",
    "Equipment Designation",
    "Equipment Type",
    "Manufacturer Sl No",
    "Voltage Ratio",
    "Cooling",
    "Manufacturing Year",
    "Weather Condition",
    "Sample ID",
    "Report Number",
    "TDCG/TGC %",
    "OST Recommendation",
    "DGA Recommendation", "Source PDF",
]

def create_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transformer Oil Data"
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr
    ws.row_dimensions[1].height = 42
    widths = [22,14,14,14,24,12,14,22,10,14,10,14,20,20,10,10,12,18,12,12,10,
              10,12,14,10,10,10,10,10,10,10,10,14,12,45,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(EXCEL_FILE)

def ev(text, pattern, default="ND"):
    try:
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else default
    except Exception:
        return default

def parse_pdf(path):
    import pdfplumber
    full = ""

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full += t + "\n"

    d = {}

    # Basic Details
    d["css_name"] = ev(full, r"Sample Location:\s*([^\n]+)", default="")
    d["transformer_no"] = ev(full, r"Equipment Serial No\s*:\s*([^\n]+)", default="")

    d["report_date"] = ev(full, r"Analysis Date\s*([0-9/]+)")
    d["sampling_date"] = ev(full, r"Sampling Date\s*:\s*([0-9/]+)")

    d["manufacturer"] = ev(full, r"Equipment Make\s*:\s*([^\n]+)", default="")
    d["rating"] = ev(full, r"Power Rating:\s*([^\n]+)", default="")
    d["voltage_class"] = ev(full, r"Voltage Rating:\s*([^\n]+)", default="")

    d["oil_type"] = ev(full, r"Product Name\s*:\s*([^\n]+)", default="")
    d["owner"] = ev(full, r"Owner\s+([^\n]+)")
    d["installation_location"] = ev(full, r"Installation Location\s+([^\n]+)")
    d["equipment_designation"] = ev(full, r"Equipment Designation\s+([^\n]+)")
    d["equipment_type"] = ev(full, r"Equipment Type\s+([^\n]+)")
    d["manufacturer_slno"] = ev(full, r"Manufacturer's Sl\. No\.\s+([^\n]+)")
    d["voltage_ratio"] = ev(full, r"Voltage Ratio.*?\s+([0-9/ ]+VOLT)")
    d["cooling"] = ev(full, r"Cooling\s+([^\n]+)")
    d["manufacturing_year"] = ev(full, r"Manufacturing Year\s+([^\n]+)")
    d["weather_condition"] = ev(full, r"Weather condition\s+([^\n]+)")
    d["sample_id"] = ev(full, r"Our Sample ID\s+([A-Z0-9\-]+)")
    d["report_no"] = ev(full, r"Oil Test Report\s*-\s*([A-Z0-9/]+)")
    # Oil Test Parameters
    d["bdv"] = ev(
        full,
        r"Electric Strength \(Breakdown Voltage\).*?([\d.]+)\s+Good"
    )

    d["water"] = ev(
        full,
        r"Water Content.*?mg/kg.*?([\d.]+)\s+Good"
    )

    d["neutralization"] = ev(
        full,
        r"Neutralization Value.*?([\d.]+)\s+Good"
    )

    d["ift"] = ev(
        full,
        r"Interfacial Tension.*?([\d.]+)\s+Good"
    )

    d["ddf_90"] = ev(
        full,
        r"Dielectric Dissipation Factor \(Tan delta\) at 900C.*?([\d.]+)\s+Good"
    )

    d["ddf_27"] = ev(
        full,
        r"Dielectric Dissipation Factor \(Tan delta\) at RT.*?([\d.]+)"
    )

    d["sp_res_90"] = ev(
        full,
        r"Specific Resistance \(Resistivity\) at 900C.*?([\d.]+)\s+Good"
    )

    d["sp_res_27"] = ev(
        full,
        r"Specific Resistance \(Resistivity\) at RT.*?([\d.]+)"
    )

    d["sediment"] = ev(
        full,
        r"Sediment & Sludge.*?(Not Detected|[\d.]+)"
    )

    d["flash"] = ev(
        full,
        r"Flash Point.*?([\d.]+)"
    )

    d["density"] = ev(
        full,
        r"Density at .*?g/ml\s*([\d.]+)"
    )

    # DGA Gases
    d["tgc"] = ev(full,
    r"Total Gas Content \(TGC\).*?([0-9.]+)\s+NS"
    )

    d["h2"] = ev(full,
        r"Hydrogen \(H₂\).*?([0-9.]+)\s+50"
    )

    d["o2"] = ev(full,
        r"Oxygen \(O₂\).*?([0-9.]+)\s+NS"
    )

    d["n2"] = ev(full,
        r"Nitrogen \(N₂\).*?([0-9.]+)\s+NS"
    )

    d["co"] = ev(full,
        r"Carbon Monoxide \(CO\).*?([0-9.]+)\s+400"
    )

    d["ch4"] = ev(full,
        r"Methane \(CH₄\).*?([0-9.]+)\s+30"
    )

    d["co2"] = ev(full,
        r"Carbon Dioxide \(CO₂\).*?([0-9.]+)\s+3800"
    )

    d["c2h4"] = ev(full,
        r"Ethylene \(C₂H₄\).*?([0-9.]+)\s+60"
    )

    d["c2h6"] = ev(full,
        r"Ethane \(C₂H₆\).*?([0-9.]+)\s+20"
    )

    d["c2h2"] = ev(full,
        r"Acetylene \(C₂H₂\).*?(ND|[0-9.]+)\s+2"
    )

    d["c3h6"] = ev(full,
        r"Propylene \(C₃H₆\).*?(ND|[0-9.]+)\s+NS"
    )

    d["c3h8"] = ev(full,
        r"Propane \(C₃H₈\).*?(ND|[0-9.]+)\s+NS"
    )

    d["tdcg"] = ev(full,
        r"Total Dissolved Combustible.*?([0-9.]+)\s+NS"
    )

    d["tdcg_ratio"] = ev(full,
        r"TDCG/TGC.*?([0-9.]+)\s+8%"
    )

    # Optional fields not present in SGS format
    d["color"] = ""
    d["oqi"] = ""
    d["c3h6"] = ""
    d["c3h8"] = ""
    d["tdcg"] = ""
    d["tgc"] = ""

    # Recommendations / comments
    d["overall_recommendation"] = ev(
        full,
        r"Overall Recommendations:\s*(.*?)\nOil Test Report"
    )

    d["ost_recommendation"] = ev(
        full,
        r"OST Test Recommendation\s*(.*?)\n#"
    )

    d["dga_recommendation"] = ev(
        full,
        r"DGA Test Recommendation\s*(.*?)\n#"
    )

    d["recommendation"] = (
        d["overall_recommendation"]
        + " | "
        + d["ost_recommendation"]
        + " | "
        + d["dga_recommendation"]
    )
    d["recommendation"] = (
    d["overall_recommendation"]
    + " | "
    + d["ost_recommendation"]
    + " | "
    + d["dga_recommendation"]
    )

    return d

def append_row(data, pdf_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    if not os.path.exists(EXCEL_FILE):
        create_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    next_row = ws.max_row + 1
    row_data = [
        data.get("owner",""),
        data.get("installation_location",""),
        data.get("equipment_designation",""),
        data.get("equipment_type",""),
        data.get("manufacturer_slno",""),
        data.get("voltage_ratio",""),
        data.get("cooling",""),
        data.get("manufacturing_year",""),
        data.get("weather_condition",""),
        data.get("sample_id",""),
        data.get("report_no",""),
        data.get("tdcg_ratio",""),
        data.get("ost_recommendation",""),
        data.get("dga_recommendation",""),
        os.path.basename(pdf_path),
    ]
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", start_color="EBF3FB") if next_row % 2 == 0 else None
    center = Alignment(horizontal="center", vertical="center")
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.border = bdr
        c.alignment = center
        c.font = Font(name="Arial", size=10)
        if fill: c.fill = fill
    ws.row_dimensions[next_row].height = 18
    wb.save(EXCEL_FILE)

def main():
    print("=" * 64)
    print("  Transformer Oil Test Report Extractor  |  TRU-FIL format")
    print(f"  Output: {EXCEL_FILE}")
    print("  Type  : 'exit' to quit")
    print("=" * 64)

    if not os.path.exists(EXCEL_FILE):
        create_excel()
        print("  Excel file created.")
    else:
        import openpyxl
        rows = openpyxl.load_workbook(EXCEL_FILE).active.max_row - 1
        print(f"  Existing file found with {rows} data row(s).")

    while True:
        print()
        choice = input("Press ENTER to select PDF or type 'exit': ").strip().lower()

        if choice == "exit":
            print(f"\n  Done. All data saved to: {EXCEL_FILE}")
            break

        Tk().withdraw()

        pdf_paths = askopenfilenames(
            title="Select Transformer Oil PDFs",
            filetypes=[("PDF Files", "*.pdf")]
        )

        if not pdf_paths:
            print("  No files selected.")
            continue

        for pdf_path in pdf_paths:

            if not os.path.exists(pdf_path):
                print(f"  ERROR: File not found -- {pdf_path}")
                continue

            if not pdf_path.lower().endswith(".pdf"):
                print(f"  Skipping non-PDF file -- {pdf_path}")
                continue

            try:
                print(f"\n  Parsing {os.path.basename(pdf_path)} ...")

                d = parse_pdf(pdf_path)
                print(type(d))
                print(d)

                append_row(d, pdf_path)

                print(f"  OK  Row appended to {EXCEL_FILE}")

                print(f"     CSS Name       : {d.get('css_name','--')}")
                print(f"     Transformer    : {d.get('transformer_no','--')}")
                print(f"     Oil Type       : {d.get('oil_type','--')}")
                print(f"     BDV (KV)       : {d.get('bdv','--')}")
                print(f"     Water (ppm)    : {d.get('water','--')}")
                print(f"     OQI            : {d.get('oqi','--')}")
                print(f"     H2/CO/C2H2     : {d.get('h2','--')} / {d.get('co','--')} / {d.get('c2h2','--')}")

            except Exception as e:
                print(f"  ERROR processing {os.path.basename(pdf_path)}")
                print(e)

if __name__ == "__main__":
    try:
        import pdfplumber, openpyxl
    except ImportError:
        print("  Installing required packages...")
        os.system(f'"{sys.executable}" -m pip install pdfplumber openpyxl --quiet')
        import pdfplumber, openpyxl
    main()
